import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import json


def create_job_profile_template():
    # Создаем новую рабочую книгу
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Профиль должности"

    # Стили
    header_font = Font(bold=True, size=12)
    section_font = Font(bold=True, size=11, color="FFFFFF")
    regular_font = Font(size=10)
    section_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_fill = PatternFill(
        start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"
    )
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    row = 1

    # Заголовок
    ws.merge_cells(f"A{row}:C{row}")
    cell = ws[f"A{row}"]
    cell.value = "ПРОФИЛЬ ДОЛЖНОСТИ"
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal="center")
    cell.fill = header_fill
    cell.border = border
    row += 2

    # Основная информация
    ws.merge_cells(f"A{row}:C{row}")
    cell = ws[f"A{row}"]
    cell.value = "ОСНОВНАЯ ИНФОРМАЦИЯ"
    cell.font = section_font
    cell.fill = section_fill
    cell.border = border
    row += 1

    basic_info = [
        ("Название должности", ""),
        ("Подразделение укрупненно", ""),
        ("Подразделение", ""),
        (
            "Категория должности",
            "Специалист/Линейный руководитель/Руководитель среднего уровня/Руководитель высшего уровня",
        ),
        ("Непосредственный руководитель", ""),
        ("Количество подразделений в подчинении", ""),
        ("Количество человек в прямом подчинении", ""),
        ("Профильная/обеспечивающая деятельность", ""),
    ]

    for label, value in basic_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = regular_font
        ws[f"A{row}"].border = border
        ws[f"B{row}"] = value
        ws[f"B{row}"].font = regular_font
        ws[f"B{row}"].border = border
        ws.merge_cells(f"B{row}:C{row}")
        row += 1

    row += 1

    # Области ответственности
    ws.merge_cells(f"A{row}:C{row}")
    cell = ws[f"A{row}"]
    cell.value = "ОБЛАСТИ ОТВЕТСТВЕННОСТИ / ДОЛЖНОСТНЫЕ ОБЯЗАННОСТИ"
    cell.font = section_font
    cell.fill = section_fill
    cell.border = border
    row += 1

    # Заголовки для областей ответственности
    ws[f"A{row}"] = "Область ответственности"
    ws[f"A{row}"].font = header_font
    ws[f"A{row}"].fill = header_fill
    ws[f"A{row}"].border = border
    ws.merge_cells(f"B{row}:C{row}")
    ws[f"B{row}"] = "Конкретные задачи"
    ws[f"B{row}"].font = header_font
    ws[f"B{row}"].fill = header_fill
    ws[f"B{row}"].border = border
    row += 1

    # Пустые строки для заполнения областей ответственности
    for i in range(5):
        ws[f"A{row}"] = ""
        ws[f"A{row}"].border = border
        ws.merge_cells(f"B{row}:C{row}")
        ws[f"B{row}"] = ""
        ws[f"B{row}"].border = border
        row += 1

    row += 1

    # Профессиональные знания и умения
    ws.merge_cells(f"A{row}:C{row}")
    cell = ws[f"A{row}"]
    cell.value = "ПРОФЕССИОНАЛЬНЫЕ ЗНАНИЯ И УМЕНИЯ"
    cell.font = section_font
    cell.fill = section_fill
    cell.border = border
    row += 1

    # Заголовки
    ws[f"A{row}"] = "Категория навыка"
    ws[f"A{row}"].font = header_font
    ws[f"A{row}"].fill = header_fill
    ws[f"A{row}"].border = border
    ws[f"B{row}"] = "Конкретные навыки"
    ws[f"B{row}"].font = header_font
    ws[f"B{row}"].fill = header_fill
    ws[f"B{row}"].border = border
    ws[f"C{row}"] = "Целевой уровень"
    ws[f"C{row}"].font = header_font
    ws[f"C{row}"].fill = header_fill
    ws[f"C{row}"].border = border
    row += 1

    # Пустые строки для навыков
    for i in range(8):
        ws[f"A{row}"] = ""
        ws[f"A{row}"].border = border
        ws[f"B{row}"] = ""
        ws[f"B{row}"].border = border
        ws[f"C{row}"] = "Базовый/Продвинутый/Экспертный" if i == 0 else ""
        ws[f"C{row}"].border = border
        row += 1

    row += 1

    # Корпоративные компетенции
    ws.merge_cells(f"A{row}:C{row}")
    cell = ws[f"A{row}"]
    cell.value = "КОРПОРАТИВНЫЕ КОМПЕТЕНЦИИ"
    cell.font = section_font
    cell.fill = section_fill
    cell.border = border
    row += 1

    competencies = [
        "Инновационность и развитие",
        "Ориентация на результат",
        "Стратегическое видение и принятие решений",
        "Клиентоориентированность",
        "Эффективная коммуникация",
        "Работа в команде",
        "Лидерство",
    ]

    for comp in competencies:
        ws[f"A{row}"] = "☐"
        ws[f"A{row}"].border = border
        ws.merge_cells(f"B{row}:C{row}")
        ws[f"B{row}"] = comp
        ws[f"B{row}"].border = border
        row += 1

    row += 1

    # Личностные качества
    ws.merge_cells(f"A{row}:C{row}")
    cell = ws[f"A{row}"]
    cell.value = "ЛИЧНОСТНЫЕ КАЧЕСТВА"
    cell.font = section_font
    cell.fill = section_fill
    cell.border = border
    row += 1

    qualities = [
        "Внимательность",
        "Ответственность",
        "Коммуникабельность",
        "Стрессоустойчивость",
        "Настойчивость",
        "Исполнительность",
        "Системность мышления",
        "Инициативность",
        "Проактивность",
        "Критическое мышление",
        "Лидерство",
        "Аналитический склад ума",
        "Многозадачность",
        "Решительность",
    ]

    for i, quality in enumerate(qualities):
        if i % 2 == 0:
            ws[f"A{row}"] = "☐"
            ws[f"A{row}"].border = border
            ws[f"B{row}"] = quality
            ws[f"B{row}"].border = border
            if i + 1 < len(qualities):
                ws[f"C{row}"] = f"☐ {qualities[i + 1]}"
                ws[f"C{row}"].border = border
            else:
                ws[f"C{row}"] = ""
                ws[f"C{row}"].border = border
            row += 1

    row += 1

    # Образование и опыт работы
    ws.merge_cells(f"A{row}:C{row}")
    cell = ws[f"A{row}"]
    cell.value = "ОБРАЗОВАНИЕ И ОПЫТ РАБОТЫ"
    cell.font = section_font
    cell.fill = section_fill
    cell.border = border
    row += 1

    education_info = [
        ("Опыт работы на предыдущей должности", ""),
        ("Общий опыт работы", ""),
        (
            "Уровень образования",
            "Среднее общее/Среднее профессиональное/Высшее/Высшее неоконченное",
        ),
        ("Специальность", ""),
        ("Дополнительное профессиональное образование", ""),
    ]

    for label, value in education_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = regular_font
        ws[f"A{row}"].border = border
        ws.merge_cells(f"B{row}:C{row}")
        ws[f"B{row}"] = value
        ws[f"B{row}"].font = regular_font
        ws[f"B{row}"].border = border
        row += 1

    row += 1

    # Карьеграмма
    ws.merge_cells(f"A{row}:C{row}")
    cell = ws[f"A{row}"]
    cell.value = "КАРЬЕГРАММА"
    cell.font = section_font
    cell.fill = section_fill
    cell.border = border
    row += 1

    ws[f"A{row}"] = "Позиции доноры"
    ws[f"A{row}"].font = header_font
    ws[f"A{row}"].fill = header_fill
    ws[f"A{row}"].border = border
    ws.merge_cells(f"B{row}:C{row}")
    ws[f"B{row}"] = ""
    ws[f"B{row}"].border = border
    row += 1

    ws[f"A{row}"] = "Целевые позиции карьерного роста"
    ws[f"A{row}"].font = header_font
    ws[f"A{row}"].fill = header_fill
    ws[f"A{row}"].border = border
    ws.merge_cells(f"B{row}:C{row}")
    ws[f"B{row}"] = ""
    ws[f"B{row}"].border = border
    row += 1

    row += 1

    # Техническое обеспечение
    ws.merge_cells(f"A{row}:C{row}")
    cell = ws[f"A{row}"]
    cell.value = "ПРОГРАММНОЕ И ТЕХНИЧЕСКОЕ ОБЕСПЕЧЕНИЕ"
    cell.font = section_font
    cell.fill = section_fill
    cell.border = border
    row += 1

    ws[f"A{row}"] = "Специальное программное обеспечение"
    ws[f"A{row}"].font = header_font
    ws[f"A{row}"].fill = header_fill
    ws[f"A{row}"].border = border
    ws.merge_cells(f"B{row}:C{row}")
    ws[f"B{row}"] = ""
    ws[f"B{row}"].border = border
    row += 1

    ws[f"A{row}"] = "Специальное оборудование"
    ws[f"A{row}"].font = header_font
    ws[f"A{row}"].fill = header_fill
    ws[f"A{row}"].border = border
    ws.merge_cells(f"B{row}:C{row}")
    ws[f"B{row}"] = ""
    ws[f"B{row}"].border = border
    row += 1

    row += 1

    # Дополнительная информация
    ws.merge_cells(f"A{row}:C{row}")
    cell = ws[f"A{row}"]
    cell.value = "ДОПОЛНИТЕЛЬНАЯ ИНФОРМАЦИЯ"
    cell.font = section_font
    cell.fill = section_fill
    cell.border = border
    row += 1

    additional_info = [("Условия повышения", ""), ("Решение о повышении", "")]

    for label, value in additional_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = regular_font
        ws[f"A{row}"].border = border
        ws.merge_cells(f"B{row}:C{row}")
        ws[f"B{row}"] = value
        ws[f"B{row}"].font = regular_font
        ws[f"B{row}"].border = border
        row += 1

    row += 1

    # Метаданные
    ws.merge_cells(f"A{row}:C{row}")
    cell = ws[f"A{row}"]
    cell.value = "МЕТАДАННЫЕ"
    cell.font = section_font
    cell.fill = section_fill
    cell.border = border
    row += 1

    metadata_info = [
        ("Подготовлено (ФИО ответственного)", ""),
        ("Дата формирования", ""),
    ]

    for label, value in metadata_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = regular_font
        ws[f"A{row}"].border = border
        ws.merge_cells(f"B{row}:C{row}")
        ws[f"B{row}"] = value
        ws[f"B{row}"].font = regular_font
        ws[f"B{row}"].border = border
        row += 1

    # Настройка ширины колонок
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 25

    # Сохранение файла
    wb.save("/home/yan/A101/HR/Profiles/job_profile_template.xlsx")
    print("XLS шаблон профиля должности создан: job_profile_template.xlsx")


if __name__ == "__main__":
    create_job_profile_template()
