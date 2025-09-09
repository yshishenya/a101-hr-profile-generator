#!/usr/bin/env python3
"""
Enhanced Excel to Markdown Converter for KPI Files
Optimized for LLM context processing with better structure detection
"""

import pandas as pd
import os
import re
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import numpy as np


class EnhancedExcelToMDConverter:
    def __init__(self, source_dir, target_dir):
        self.source_dir = Path(source_dir)
        self.target_dir = Path(target_dir)
        self.target_dir.mkdir(exist_ok=True)

    def clean_text(self, text):
        """Clean and normalize text for markdown"""
        if pd.isna(text) or text is None:
            return ""
        text = str(text).strip()
        # Remove extra spaces and normalize line breaks
        text = re.sub(r"\s+", " ", text)
        text = re.sub(r"\n+", "\n", text)
        return text

    def extract_department_info(self, filename):
        """Extract department and employee info from filename"""
        # Parse filename pattern: КПЭ 2025_DEPT (NAME)+.xlsx
        match = re.match(r"КПЭ 2025_(.+?)(?:\s*\((.+?)\))?[+_]?\.xlsx?$", filename)
        if match:
            dept = match.group(1).strip()
            employee = match.group(2).strip() if match.group(2) else ""
            return dept, employee
        return filename.replace(".xlsx", "").replace(".xls", ""), ""

    def detect_kpi_sections(self, rows_data):
        """Detect KPI sections in the data"""
        sections = []
        current_section = None

        for i, row in enumerate(rows_data):
            row_text = " ".join(
                [str(cell) for cell in row if cell and str(cell).strip()]
            )

            # Look for section headers
            if any(
                keyword in row_text.lower()
                for keyword in [
                    "корпоративные кпэ",
                    "личные кпэ",
                    "бюджетные кпэ",
                    "кпэ",
                    "показател",
                ]
            ):
                if current_section:
                    sections.append(current_section)
                current_section = {"name": row_text, "start_row": i, "data": []}
            elif current_section:
                current_section["data"].append((i, row))

        if current_section:
            sections.append(current_section)

        return sections

    def parse_kpi_table(self, data_rows):
        """Parse KPI table data into structured format"""
        kpi_items = []
        headers = []

        for i, (row_idx, row) in enumerate(data_rows):
            non_empty_cells = [
                self.clean_text(cell) for cell in row if self.clean_text(cell)
            ]

            if not non_empty_cells:
                continue

            # First meaningful row might be headers
            if i == 0 or (
                len(non_empty_cells) >= 3
                and any(
                    keyword in " ".join(non_empty_cells).lower()
                    for keyword in ["кпэ", "значение", "ед.изм", "показател", "цел"]
                )
            ):
                headers = non_empty_cells
                continue

            # Parse KPI data rows
            if len(non_empty_cells) >= 2:
                kpi_item = {
                    "name": non_empty_cells[0],
                    "target_value": (
                        non_empty_cells[1] if len(non_empty_cells) > 1 else ""
                    ),
                    "unit": non_empty_cells[2] if len(non_empty_cells) > 2 else "",
                    "weights": non_empty_cells[3:] if len(non_empty_cells) > 3 else [],
                    "description": "",
                }

                # Try to extract more structured data
                if len(headers) > 3:
                    for j, header in enumerate(headers[3:], start=3):
                        if j < len(non_empty_cells):
                            if any(
                                keyword in header.lower()
                                for keyword in ["методика", "описание", "источник"]
                            ):
                                kpi_item[
                                    "description"
                                ] += f"{header}: {non_empty_cells[j]}; "

                kpi_items.append(kpi_item)

        return kpi_items, headers

    def convert_sheet_to_structured_md(
        self, worksheet, sheet_name, dept_name, employee_name
    ):
        """Convert worksheet to well-structured markdown optimized for LLM"""
        # Extract all data
        rows_data = []
        for row in worksheet.iter_rows(values_only=True):
            rows_data.append([self.clean_text(cell) for cell in row])

        # Detect KPI sections
        sections = self.detect_kpi_sections(rows_data)

        md_content = []

        # Document header
        md_content.append(f"# KPI документ: {dept_name}")
        md_content.append("")

        # Metadata section
        md_content.append("## Общая информация")
        md_content.append("")
        if dept_name:
            md_content.append(f"- **Департамент/Отдел:** {dept_name}")
        if employee_name:
            md_content.append(f"- **Ответственный сотрудник:** {employee_name}")
        if sheet_name and sheet_name != "Заполнить":
            md_content.append(f"- **Лист:** {sheet_name}")
        md_content.append("")

        # Process each detected section
        if sections:
            for section in sections:
                md_content.append(f"## {section['name']}")
                md_content.append("")

                # Parse KPI items in this section
                kpi_items, headers = self.parse_kpi_table(section["data"])

                if kpi_items:
                    for kpi in kpi_items:
                        md_content.append(f"### {kpi['name']}")
                        md_content.append("")

                        if kpi["target_value"]:
                            md_content.append(
                                f"- **Целевое значение:** {kpi['target_value']}"
                            )
                        if kpi["unit"]:
                            md_content.append(f"- **Единица измерения:** {kpi['unit']}")
                        if kpi["description"]:
                            md_content.append(
                                f"- **Дополнительная информация:** {kpi['description'].strip('; ')}"
                            )

                        if kpi["weights"]:
                            md_content.append("- **Веса/Коэффициенты:**")
                            for i, weight in enumerate(kpi["weights"]):
                                if weight and str(weight).strip():
                                    header_name = (
                                        headers[i + 3]
                                        if i + 3 < len(headers)
                                        else f"Позиция {i+1}"
                                    )
                                    md_content.append(f"  - {header_name}: {weight}")

                        md_content.append("")
                else:
                    # Fallback: show raw table data
                    md_content.append("| Показатель | Значение | Единица измерения |")
                    md_content.append("| --- | --- | --- |")

                    for _, row in section["data"]:
                        non_empty = [
                            self.clean_text(cell)
                            for cell in row
                            if self.clean_text(cell)
                        ]
                        if len(non_empty) >= 2:
                            row_md = "| " + " | ".join(non_empty[:3])
                            while len(non_empty) < 3:
                                row_md += " |"
                            if len(non_empty) == 2:
                                row_md += " |"
                            md_content.append(row_md)
                    md_content.append("")
        else:
            # Fallback: process all data as single table
            md_content.append("## Данные KPI")
            md_content.append("")

            for i, row in enumerate(rows_data):
                non_empty = [cell for cell in row if cell and str(cell).strip()]
                if len(non_empty) >= 2:
                    if i == 0 or any(
                        keyword in " ".join(non_empty).lower()
                        for keyword in ["кпэ", "показател", "цел"]
                    ):
                        # Header row
                        md_content.append("| " + " | ".join(non_empty) + " |")
                        md_content.append(
                            "| " + " | ".join(["---"] * len(non_empty)) + " |"
                        )
                    else:
                        # Data row
                        md_content.append("| " + " | ".join(non_empty) + " |")

        return "\n".join(md_content)

    def convert_file(self, excel_file):
        """Convert a single Excel file to structured markdown"""
        print(f"Converting {excel_file.name}...")

        dept_name, employee_name = self.extract_department_info(excel_file.name)

        try:
            # Load workbook
            workbook = load_workbook(excel_file, read_only=True, data_only=True)

            # Get the main worksheet
            main_sheet = None
            max_data_rows = 0

            for sheet_name in workbook.sheetnames:
                ws = workbook[sheet_name]
                data_rows = sum(
                    1 for row in ws.iter_rows() if any(cell.value for cell in row)
                )
                if data_rows > max_data_rows:
                    max_data_rows = data_rows
                    main_sheet = ws

            if main_sheet is None:
                print(f"  Warning: No data found in {excel_file.name}")
                return

            # Convert to structured markdown
            md_content = self.convert_sheet_to_structured_md(
                main_sheet, main_sheet.title, dept_name, employee_name
            )

            # Save markdown file
            md_filename = excel_file.stem + "_structured.md"
            md_file_path = self.target_dir / md_filename

            with open(md_file_path, "w", encoding="utf-8") as f:
                f.write(md_content)

            print(f"  ✓ Converted to {md_file_path}")

        except Exception as e:
            print(f"  ✗ Error converting {excel_file.name}: {e}")

    def convert_all_files(self):
        """Convert all Excel files in the source directory"""
        excel_files = list(self.source_dir.glob("*.xlsx")) + list(
            self.source_dir.glob("*.xls")
        )

        if not excel_files:
            print("No Excel files found in source directory")
            return

        print(f"Found {len(excel_files)} Excel files to convert:")
        for file in excel_files:
            print(f"  - {file.name}")
        print()

        for excel_file in excel_files:
            self.convert_file(excel_file)

        print(f"\nEnhanced conversion completed! {len(excel_files)} files processed.")
        print(f"Structured markdown files saved to: {self.target_dir}")


def main():
    source_dir = "/home/yan/A101/HR/KPI/"
    target_dir = "/home/yan/A101/HR/KPI/md_converted/"

    converter = EnhancedExcelToMDConverter(source_dir, target_dir)
    converter.convert_all_files()


if __name__ == "__main__":
    main()
