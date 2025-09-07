#!/usr/bin/env python3
"""
Анализатор структуры KPI Excel файлов
Анализирует сложность структуры данных и форматирования
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border
import os
import json
from collections import defaultdict

def analyze_excel_structure(file_path):
    """Детальный анализ структуры Excel файла"""
    print(f"\n{'='*60}")
    print(f"АНАЛИЗ ФАЙЛА: {os.path.basename(file_path)}")
    print(f"{'='*60}")
    
    analysis = {
        'filename': os.path.basename(file_path),
        'file_size_kb': round(os.path.getsize(file_path) / 1024, 2),
        'sheets': {},
        'complexity_indicators': {
            'merged_cells_count': 0,
            'different_fonts_count': 0,
            'different_fills_count': 0,
            'formulas_count': 0,
            'empty_cells_in_data': 0,
            'inconsistent_data_types': False,
            'complex_headers': False,
            'nested_structure': False
        },
        'recommendations': []
    }
    
    try:
        # Загружаем workbook для детального анализа форматирования
        wb = openpyxl.load_workbook(file_path, data_only=False)
        
        print(f"Количество листов: {len(wb.sheetnames)}")
        print(f"Названия листов: {wb.sheetnames}")
        
        for sheet_name in wb.sheetnames:
            print(f"\n--- ЛИСТ: {sheet_name} ---")
            ws = wb[sheet_name]
            
            sheet_analysis = {
                'name': sheet_name,
                'dimensions': f"{ws.min_row}:{ws.max_row} x {ws.min_column}:{ws.max_column}",
                'used_range': f"{ws.max_row - ws.min_row + 1} строк x {ws.max_column - ws.min_column + 1} столбцов",
                'merged_cells': [],
                'data_preview': [],
                'formatting_complexity': {},
                'data_types_found': set()
            }
            
            # Анализ объединенных ячеек
            merged_ranges = list(ws.merged_cells.ranges)
            sheet_analysis['merged_cells'] = [str(mr) for mr in merged_ranges]
            analysis['complexity_indicators']['merged_cells_count'] += len(merged_ranges)
            
            print(f"  Размеры: {sheet_analysis['used_range']}")
            print(f"  Объединенные ячейки: {len(merged_ranges)}")
            if merged_ranges:
                print(f"    Примеры: {merged_ranges[:3]}")
            
            # Анализ форматирования и данных
            fonts = set()
            fills = set()
            formulas = 0
            empty_in_data = 0
            
            # Получаем превью данных (первые 10 строк и столбцов)
            preview_data = []
            for row in range(1, min(11, ws.max_row + 1)):
                row_data = []
                for col in range(1, min(11, ws.max_column + 1)):
                    cell = ws.cell(row=row, column=col)
                    
                    # Сбор статистики по форматированию
                    if cell.font:
                        fonts.add(str(cell.font))
                    if cell.fill and cell.fill.start_color and cell.fill.start_color.index:
                        fills.add(str(cell.fill))
                    
                    # Подсчет формул
                    if cell.data_type == 'f':
                        formulas += 1
                    
                    # Анализ типов данных
                    if cell.value is not None:
                        sheet_analysis['data_types_found'].add(type(cell.value).__name__)
                        row_data.append(str(cell.value)[:50])  # Ограничиваем длину
                    else:
                        empty_in_data += 1
                        row_data.append('')
                
                preview_data.append(row_data)
            
            sheet_analysis['data_preview'] = preview_data
            sheet_analysis['data_types_found'] = list(sheet_analysis['data_types_found'])
            
            analysis['complexity_indicators']['different_fonts_count'] += len(fonts)
            analysis['complexity_indicators']['different_fills_count'] += len(fills)
            analysis['complexity_indicators']['formulas_count'] += formulas
            analysis['complexity_indicators']['empty_cells_in_data'] += empty_in_data
            
            print(f"  Различных шрифтов: {len(fonts)}")
            print(f"  Различных заливок: {len(fills)}")
            print(f"  Формул: {formulas}")
            print(f"  Пустых ячеек в данных: {empty_in_data}")
            print(f"  Типы данных: {sheet_analysis['data_types_found']}")
            
            # Показываем превью данных
            print(f"\n  ПРЕВЬЮ ДАННЫХ (первые строки и столбцы):")
            for i, row in enumerate(preview_data[:5]):
                print(f"    Строка {i+1}: {row[:5]}")
            
            analysis['sheets'][sheet_name] = sheet_analysis
        
        # Попытка загрузить данные через pandas для дополнительного анализа
        print(f"\n--- АНАЛИЗ ЧЕРЕЗ PANDAS ---")
        try:
            # Читаем все листы через pandas
            excel_data = pd.read_excel(file_path, sheet_name=None, header=None)
            
            for sheet_name, df in excel_data.items():
                print(f"\nЛист '{sheet_name}' через pandas:")
                print(f"  Размер DataFrame: {df.shape}")
                print(f"  NaN значений: {df.isna().sum().sum()}")
                
                # Анализ структуры заголовков
                header_complexity = analyze_header_structure(df)
                if header_complexity['is_complex']:
                    analysis['complexity_indicators']['complex_headers'] = True
                
                print(f"  Сложность заголовков: {'Высокая' if header_complexity['is_complex'] else 'Низкая'}")
                if header_complexity['issues']:
                    print(f"    Проблемы: {header_complexity['issues']}")
        
        except Exception as e:
            print(f"Ошибка при анализе через pandas: {e}")
    
    except Exception as e:
        print(f"Ошибка при анализе файла: {e}")
        analysis['error'] = str(e)
    
    # Определение общей сложности и рекомендаций
    complexity_score = calculate_complexity_score(analysis['complexity_indicators'])
    analysis['complexity_score'] = complexity_score
    analysis['complexity_level'] = get_complexity_level(complexity_score)
    analysis['recommendations'] = generate_recommendations(analysis)
    
    print(f"\n{'='*40}")
    print(f"ИТОГОВАЯ ОЦЕНКА СЛОЖНОСТИ: {complexity_score}/100")
    print(f"УРОВЕНЬ СЛОЖНОСТИ: {analysis['complexity_level']}")
    print(f"{'='*40}")
    
    return analysis

def analyze_header_structure(df):
    """Анализ структуры заголовков"""
    result = {
        'is_complex': False,
        'issues': []
    }
    
    # Проверяем первые несколько строк на предмет сложной структуры заголовков
    if df.shape[0] > 0:
        first_rows = df.head(5)
        
        # Ищем признаки многоуровневых заголовков
        for i in range(min(3, len(first_rows))):
            row = first_rows.iloc[i]
            non_null_count = row.count()
            
            # Если в строке много пустых ячеек между заполненными - возможно многоуровневый заголовок
            if non_null_count > 0 and non_null_count < len(row) * 0.7:
                result['is_complex'] = True
                result['issues'].append(f"Возможен многоуровневый заголовок в строке {i+1}")
    
    return result

def calculate_complexity_score(indicators):
    """Расчет общего балла сложности (0-100)"""
    score = 0
    
    # Объединенные ячейки (0-30 баллов)
    merged_penalty = min(indicators['merged_cells_count'] * 2, 30)
    score += merged_penalty
    
    # Разнообразие форматирования (0-20 баллов)
    format_penalty = min((indicators['different_fonts_count'] + indicators['different_fills_count']) * 2, 20)
    score += format_penalty
    
    # Формулы (0-15 баллов)
    formula_penalty = min(indicators['formulas_count'] * 3, 15)
    score += formula_penalty
    
    # Сложные заголовки (0-20 баллов)
    if indicators['complex_headers']:
        score += 20
    
    # Вложенная структура (0-15 баллов)
    if indicators['nested_structure']:
        score += 15
    
    return min(score, 100)

def get_complexity_level(score):
    """Определение уровня сложности по баллам"""
    if score <= 20:
        return "НИЗКАЯ - подходит простой парсер"
    elif score <= 50:
        return "СРЕДНЯЯ - нужен продвинутый парсер"
    else:
        return "ВЫСОКАЯ - рекомендуется LLM"

def generate_recommendations(analysis):
    """Генерация рекомендаций на основе анализа"""
    recommendations = []
    indicators = analysis['complexity_indicators']
    
    if indicators['merged_cells_count'] > 10:
        recommendations.append("Много объединенных ячеек - нужна специальная обработка структуры")
    
    if indicators['complex_headers']:
        recommendations.append("Сложная структура заголовков - потребуется интеллектуальный парсинг")
    
    if indicators['formulas_count'] > 5:
        recommendations.append("Присутствуют формулы - нужно решить, использовать значения или формулы")
    
    if indicators['different_fonts_count'] > 3:
        recommendations.append("Разнообразное форматирование может содержать семантическую информацию")
    
    complexity_score = analysis['complexity_score']
    if complexity_score <= 20:
        recommendations.append("✅ РЕКОМЕНДАЦИЯ: Достаточно простого Excel парсера (pandas/openpyxl)")
    elif complexity_score <= 50:
        recommendations.append("⚠️ РЕКОМЕНДАЦИЯ: Нужен продвинутый парсер с обработкой структуры")
    else:
        recommendations.append("🤖 РЕКОМЕНДАЦИЯ: Использовать LLM для интерпретации сложной структуры")
    
    return recommendations

def main():
    """Основная функция для анализа указанных файлов"""
    kpi_dir = "/home/yan/A101/HR/KPI"
    target_files = [
        "КПЭ 2025_АС (Ларина О)+.xlsx",
        "КПЭ 2025_ДИТ (Сложеникин А)+.xlsx", 
        "КПЭ 2025_ТОП_финал_.xlsx"
    ]
    
    all_analyses = []
    
    print("АНАЛИЗ СТРУКТУРЫ KPI ФАЙЛОВ")
    print("="*60)
    
    for filename in target_files:
        file_path = os.path.join(kpi_dir, filename)
        if os.path.exists(file_path):
            analysis = analyze_excel_structure(file_path)
            all_analyses.append(analysis)
        else:
            print(f"ФАЙЛ НЕ НАЙДЕН: {filename}")
    
    # Общий анализ и рекомендации
    print(f"\n\n{'='*80}")
    print("ОБЩИЙ АНАЛИЗ И РЕКОМЕНДАЦИИ")
    print(f"{'='*80}")
    
    if all_analyses:
        avg_complexity = sum(a['complexity_score'] for a in all_analyses) / len(all_analyses)
        max_complexity = max(a['complexity_score'] for a in all_analyses)
        
        print(f"Средняя сложность: {avg_complexity:.1f}/100")
        print(f"Максимальная сложность: {max_complexity}/100")
        
        total_merged_cells = sum(a['complexity_indicators']['merged_cells_count'] for a in all_analyses)
        total_formulas = sum(a['complexity_indicators']['formulas_count'] for a in all_analyses)
        
        print(f"Всего объединенных ячеек: {total_merged_cells}")
        print(f"Всего формул: {total_formulas}")
        
        # Финальная рекомендация
        print(f"\n{'='*50}")
        print("🎯 ФИНАЛЬНАЯ РЕКОМЕНДАЦИЯ:")
        print(f"{'='*50}")
        
        if max_complexity <= 20:
            print("✅ Простой Excel парсер (pandas) - ДОСТАТОЧНО")
            print("   - Структура данных относительно простая")
            print("   - Минимум объединенных ячеек и сложного форматирования")
        elif max_complexity <= 50:
            print("⚠️ Продвинутый парсер - РЕКОМЕНДУЕТСЯ")
            print("   - Требуется обработка объединенных ячеек")
            print("   - Нужна логика для интерпретации структуры")
        else:
            print("🤖 LLM для парсинга - НЕОБХОДИМ")
            print("   - Высокая сложность структуры")
            print("   - Много объединенных ячеек и сложного форматирования")
            print("   - Требуется интеллектуальная интерпретация контента")
    
    # Сохраняем результаты анализа
    output_path = "/home/yan/A101/HR/kpi_analysis_results.json"
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(all_analyses, f, ensure_ascii=False, indent=2)
    
    print(f"\n📄 Детальные результаты сохранены в: {output_path}")

if __name__ == "__main__":
    main()