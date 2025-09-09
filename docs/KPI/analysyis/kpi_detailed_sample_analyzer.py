#!/usr/bin/env python3
"""
Детальный анализ образцов данных из KPI файлов
"""

import pandas as pd
import openpyxl
import os


def analyze_data_samples():
    """Показывает реальные примеры данных из файлов"""
    kpi_dir = "/home/yan/A101/HR/KPI"
    target_files = [
        "КПЭ 2025_АС (Ларина О)+.xlsx",
        "КПЭ 2025_ДИТ (Сложеникин А)+.xlsx",
        "КПЭ 2025_ТОП_финал_.xlsx",
    ]

    print("ДЕТАЛЬНЫЙ АНАЛИЗ ПРИМЕРОВ ДАННЫХ KPI")
    print("=" * 60)

    for filename in target_files:
        file_path = os.path.join(kpi_dir, filename)
        if not os.path.exists(file_path):
            continue

        print(f"\n📁 ФАЙЛ: {filename}")
        print("-" * 50)

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            main_sheet = wb.worksheets[0]  # Основной лист

            print(f"Основной лист: '{main_sheet.title}'")
            print(
                f"Размер: {main_sheet.max_row} строк x {main_sheet.max_column} столбцов"
            )

            # Показываем структуру заголовков
            print(f"\n🏷️ СТРУКТУРА ЗАГОЛОВКОВ:")
            for row_num in range(1, min(6, main_sheet.max_row + 1)):
                row_values = []
                for col_num in range(1, min(16, main_sheet.max_column + 1)):
                    cell = main_sheet.cell(row=row_num, column=col_num)
                    value = str(cell.value) if cell.value else ""
                    if len(value) > 30:
                        value = value[:27] + "..."
                    row_values.append(value)

                # Убираем пустые значения в конце
                while row_values and not row_values[-1]:
                    row_values.pop()

                if row_values:
                    print(f"  Строка {row_num}: {row_values}")

            # Показываем примеры данных KPI
            print(f"\n📊 ПРИМЕРЫ KPI ДАННЫХ:")
            kpi_data_found = False

            for row_num in range(5, min(15, main_sheet.max_row + 1)):
                row_values = []
                has_meaningful_data = False

                for col_num in range(1, min(8, main_sheet.max_column + 1)):
                    cell = main_sheet.cell(row=row_num, column=col_num)
                    value = cell.value

                    if value is not None and str(value).strip():
                        has_meaningful_data = True
                        if isinstance(value, str) and len(value) > 40:
                            value = value[:37] + "..."
                        row_values.append(str(value))
                    else:
                        row_values.append("")

                if has_meaningful_data and len([v for v in row_values if v]) >= 2:
                    kpi_data_found = True
                    print(f"  KPI {row_num-4}: {row_values}")

                    if row_num - 4 >= 5:  # Ограничиваем количество примеров
                        break

            if not kpi_data_found:
                print("  Структурированные KPI данные не найдены в ожидаемом диапазоне")

            # Анализ объединенных ячеек
            merged_ranges = list(main_sheet.merged_cells.ranges)
            if merged_ranges:
                print(f"\n🔗 ОБЪЕДИНЕННЫЕ ЯЧЕЙКИ ({len(merged_ranges)}):")
                for i, mr in enumerate(merged_ranges[:8]):  # Показываем первые 8
                    # Получаем значение из первой ячейки диапазона
                    top_left = main_sheet.cell(mr.min_row, mr.min_col)
                    value = (
                        str(top_left.value)[:30] + "..."
                        if top_left.value and len(str(top_left.value)) > 30
                        else str(top_left.value)
                    )
                    print(f"  {mr}: '{value}'")

                if len(merged_ranges) > 8:
                    print(f"  ... и еще {len(merged_ranges) - 8} объединенных областей")

        except Exception as e:
            print(f"❌ Ошибка при анализе файла: {e}")

    # Анализ через pandas для понимания структуры данных
    print(f"\n\n📈 АНАЛИЗ ЧЕРЕЗ PANDAS (структурированные данные)")
    print("=" * 60)

    for filename in target_files:
        file_path = os.path.join(kpi_dir, filename)
        if not os.path.exists(file_path):
            continue

        print(f"\n📁 {filename}")
        print("-" * 30)

        try:
            # Пробуем разные способы чтения
            df = pd.read_excel(file_path, header=None)
            print(f"Общий размер через pandas: {df.shape}")

            # Ищем возможные области данных
            print("Поиск структурированных областей данных:")

            # Ищем строки с KPI (обычно содержат ключевые слова)
            kpi_keywords = ["КПЭ", "KPI", "показатель", "метрика", "цел"]

            for idx, row in df.iterrows():
                if idx > 20:  # Не ищем слишком далеко
                    break

                row_str = " ".join([str(val) for val in row.values if val is not None])
                if any(keyword.lower() in row_str.lower() for keyword in kpi_keywords):
                    print(
                        f"  Строка {idx+1} (возможный заголовок): {row.values[:6].tolist()}"
                    )

            # Показываем области с числовыми данными
            numeric_areas = []
            for idx, row in df.iterrows():
                if idx > 30:
                    break
                numeric_count = sum(
                    1
                    for val in row.values
                    if isinstance(val, (int, float)) and not pd.isna(val)
                )
                if numeric_count >= 2:
                    numeric_areas.append((idx, numeric_count))

            if numeric_areas:
                print(f"Области с числовыми данными:")
                for row_idx, count in numeric_areas[:5]:
                    print(f"  Строка {row_idx+1}: {count} числовых значений")

        except Exception as e:
            print(f"❌ Ошибка pandas анализа: {e}")


if __name__ == "__main__":
    analyze_data_samples()
