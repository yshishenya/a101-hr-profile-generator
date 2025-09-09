#!/usr/bin/env python3
"""
Excel to Markdown Converter for KPI Files
Converts Excel files with complex structures to LLM-optimized markdown format
"""

import pandas as pd
import os
import re
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import numpy as np


class ExcelToMDConverter:
    def __init__(self, source_dir, target_dir):
        self.source_dir = Path(source_dir)
        self.target_dir = Path(target_dir)
        self.target_dir.mkdir(exist_ok=True)

    def clean_text(self, text):
        """Clean and normalize text for markdown"""
        if pd.isna(text) or text is None:
            return ""
        text = str(text).strip()
        # Remove extra spaces and normalize line breaks
        text = re.sub(r"\s+", " ", text)
        text = re.sub(r"\n+", "\n", text)
        return text

    def extract_department_info(self, filename):
        """Extract department and employee info from filename"""
        # Parse filename pattern: КПЭ 2025_DEPT (NAME)+.xlsx
        match = re.match(r"КПЭ 2025_(.+?)(?:\s*\((.+?)\))?[+_]?\.xlsx?$", filename)
        if match:
            dept = match.group(1).strip()
            employee = match.group(2).strip() if match.group(2) else ""
            return dept, employee
        return filename.replace(".xlsx", "").replace(".xls", ""), ""

    def analyze_sheet_structure(self, worksheet):
        """Analyze the structure of the worksheet to identify KPI sections"""
        rows_data = []
        for row in worksheet.iter_rows(values_only=True):
            rows_data.append([self.clean_text(cell) for cell in row])

        # Find header patterns and structure
        structure = {"headers": [], "kpi_sections": [], "data_rows": []}

        for i, row in enumerate(rows_data):
            row_text = " ".join([str(cell) for cell in row if cell])

            # Look for KPI section headers
            if any(
                keyword in row_text.lower()
                for keyword in ["kpi", "кпэ", "показател", "цел", "критери"]
            ):
                structure["headers"].append((i, row))

            # Look for data rows (non-empty rows with structured data)
            non_empty_cells = [cell for cell in row if cell and str(cell).strip()]
            if len(non_empty_cells) >= 2:
                structure["data_rows"].append((i, row))

        return structure, rows_data

    def convert_sheet_to_md(self, worksheet, sheet_name, dept_name, employee_name):
        """Convert a worksheet to markdown format"""
        structure, rows_data = self.analyze_sheet_structure(worksheet)

        md_content = []

        # Header section
        md_content.append(f"# KPI {sheet_name}")
        md_content.append("")

        if dept_name:
            md_content.append(f"**Департамент/Отдел:** {dept_name}")
        if employee_name:
            md_content.append(f"**Сотрудник:** {employee_name}")
        md_content.append("")

        # Process the data
        current_section = None
        in_table = False
        table_headers = []

        for i, row in enumerate(rows_data):
            row_text = " ".join(
                [str(cell) for cell in row if cell and str(cell).strip()]
            )

            if not row_text.strip():
                if in_table:
                    md_content.append("")
                    in_table = False
                continue

            # Check if this row looks like a section header
            if len([cell for cell in row if cell and str(cell).strip()]) <= 2 and any(
                keyword in row_text.lower()
                for keyword in ["kpi", "кпэ", "показател", "корпоратив", "личн", "цел"]
            ):

                if in_table:
                    md_content.append("")
                    in_table = False

                md_content.append(f"## {row_text}")
                md_content.append("")
                current_section = row_text
                continue

            # Check if this row looks like table headers
            non_empty_cells = [
                str(cell).strip() for cell in row if cell and str(cell).strip()
            ]
            if len(non_empty_cells) >= 3 and any(
                keyword in " ".join(non_empty_cells).lower()
                for keyword in [
                    "показател",
                    "значени",
                    "описан",
                    "вес",
                    "цел",
                    "критери",
                    "ед.изм",
                ]
            ):

                table_headers = non_empty_cells
                md_content.append("| " + " | ".join(table_headers) + " |")
                md_content.append(
                    "| " + " | ".join(["---"] * len(table_headers)) + " |"
                )
                in_table = True
                continue

            # Regular data row
            if len(non_empty_cells) >= 2:
                if in_table and table_headers:
                    # Ensure we have the right number of columns
                    while len(non_empty_cells) < len(table_headers):
                        non_empty_cells.append("")
                    row_data = non_empty_cells[: len(table_headers)]
                    md_content.append("| " + " | ".join(row_data) + " |")
                else:
                    # Not in a table, format as bullet points or paragraphs
                    if len(non_empty_cells) == 1:
                        md_content.append(f"- {non_empty_cells[0]}")
                    else:
                        md_content.append(
                            f"**{non_empty_cells[0]}:** {' | '.join(non_empty_cells[1:])}"
                        )

        return "\n".join(md_content)

    def convert_file(self, excel_file):
        """Convert a single Excel file to markdown"""
        print(f"Converting {excel_file.name}...")

        dept_name, employee_name = self.extract_department_info(excel_file.name)

        try:
            # Load workbook
            workbook = load_workbook(excel_file, read_only=True, data_only=True)

            # Get the main worksheet (usually the first one or one with most data)
            main_sheet = None
            max_data_rows = 0

            for sheet_name in workbook.sheetnames:
                ws = workbook[sheet_name]
                data_rows = sum(
                    1 for row in ws.iter_rows() if any(cell.value for cell in row)
                )
                if data_rows > max_data_rows:
                    max_data_rows = data_rows
                    main_sheet = ws

            if main_sheet is None:
                print(f"  Warning: No data found in {excel_file.name}")
                return

            # Convert to markdown
            md_content = self.convert_sheet_to_md(
                main_sheet, main_sheet.title, dept_name, employee_name
            )

            # Save markdown file
            md_filename = excel_file.stem + ".md"
            md_file_path = self.target_dir / md_filename

            with open(md_file_path, "w", encoding="utf-8") as f:
                f.write(md_content)

            print(f"  ✓ Converted to {md_file_path}")

        except Exception as e:
            print(f"  ✗ Error converting {excel_file.name}: {e}")

    def convert_all_files(self):
        """Convert all Excel files in the source directory"""
        excel_files = list(self.source_dir.glob("*.xlsx")) + list(
            self.source_dir.glob("*.xls")
        )

        if not excel_files:
            print("No Excel files found in source directory")
            return

        print(f"Found {len(excel_files)} Excel files to convert:")
        for file in excel_files:
            print(f"  - {file.name}")
        print()

        for excel_file in excel_files:
            self.convert_file(excel_file)

        print(f"\nConversion completed! {len(excel_files)} files processed.")
        print(f"Markdown files saved to: {self.target_dir}")


def main():
    source_dir = "/home/yan/A101/HR/KPI/"
    target_dir = "/home/yan/A101/HR/KPI/md_converted/"

    converter = ExcelToMDConverter(source_dir, target_dir)
    converter.convert_all_files()


if __name__ == "__main__":
    main()
