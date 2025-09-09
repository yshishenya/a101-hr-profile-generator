#!/usr/bin/env python3
"""
Утилита для безопасного дампа структуры XLSX (все листы, строки и значения)
в JSON для последующего построения JSON Schema.

Пример:
  python3 backend/tools/xlsx_dump.py \
    --input "/home/yan/A101/HR/docs/Profiles/Профили архитекторы.xlsx" \
    --output "/home/yan/A101/HR/tmp/architects_xlsx_dump.json" \
    --max-rows 600 --max-cols 40
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any, Dict, List


def _to_serializable(value: Any) -> Any:
    """Конвертирует значения ячеек в сериализуемый JSON формат."""
    if value is None:
        return None
    if isinstance(value, (str, int, float, bool)):
        # Строки дополнительно триммим
        return value.strip() if isinstance(value, str) else value
    # Остальные типы (даты и пр.) приводим к строке
    try:
        return str(value)
    except Exception:
        return repr(value)


def dump_workbook(input_path: Path, max_rows: int, max_cols: int) -> Dict[str, Any]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as e:
        raise SystemExit(
            "Не найден openpyxl. Установите: python3 -m pip install --user openpyxl"
        ) from e

    if not input_path.exists():
        raise SystemExit(f"XLSX файл не найден: {input_path}")

    wb = load_workbook(filename=str(input_path), data_only=True, read_only=True)

    result: Dict[str, Any] = {
        "workbook": str(input_path),
        "sheets": []
    }

    for ws in wb.worksheets:
        sheet_info: Dict[str, Any] = {
            "name": ws.title,
            "dimensions": {
                "max_row": ws.max_row,
                "max_column": ws.max_column
            },
            "merged": [],
            "rows": []
        }

        # Сведения о merged-ячейках (полезно для заголовков-секций)
        try:
            sheet_info["merged"] = [str(rng) for rng in ws.merged_cells.ranges]
        except Exception:
            sheet_info["merged"] = []

        # Ограничиваем объём, но достаточно большой, чтобы покрыть формы
        scan_rows = min(ws.max_row or 0, max_rows)
        scan_cols = min(ws.max_column or 0, max_cols)

        for r in range(1, scan_rows + 1):
            row_vals: List[Any] = []
            non_empty = False
            for c in range(1, scan_cols + 1):
                v = ws.cell(r, c).value
                v = _to_serializable(v)
                if v not in (None, ""):
                    non_empty = True
                row_vals.append(v)
            if non_empty:
                sheet_info["rows"].append({
                    "r": r,
                    "values": row_vals
                })

        result["sheets"].append(sheet_info)

    return result


def main() -> None:
    parser = argparse.ArgumentParser(description="Дамп структуры XLSX в JSON")
    parser.add_argument("--input", required=True, help="Путь к XLSX файлу")
    parser.add_argument("--output", required=True, help="Путь к JSON дампу")
    parser.add_argument("--max-rows", type=int, default=600, help="Максимум строк на лист")
    parser.add_argument("--max-cols", type=int, default=40, help="Максимум столбцов на лист")

    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    data = dump_workbook(input_path=input_path, max_rows=args.max_rows, max_cols=args.max_cols)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(json.dumps({
        "ok": True,
        "input": str(input_path),
        "output": str(output_path),
        "sheets": [s["name"] for s in data["sheets"]]
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()




