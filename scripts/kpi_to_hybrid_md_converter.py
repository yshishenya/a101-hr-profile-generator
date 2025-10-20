#!/usr/bin/env python3
"""
KPI to Hybrid Markdown Converter
Converts Excel KPI files to Hybrid MD format (YAML frontmatter + MD table)
Solves the ambiguous column problem for LLM-based profile generation

Features:
- YAML frontmatter with department metadata
- Clarified position columns using employee names
- Clean markdown table format
- Optimized for LLM context
"""

import pandas as pd
import yaml
import re
from pathlib import Path
from openpyxl import load_workbook
from typing import Dict, List, Tuple, Optional


class KPIToHybridMDConverter:
    """
    Converter that creates Hybrid MD format:
    ---
    department: ДИТ
    responsible: Сложеникин А
    positions_map:
      Директор по ИТ: Сложеникин А.В.
    ---
    | КПЭ | ... | Директор (Сложеникин) | ...
    """

    def __init__(self, source_dir: str, target_dir: str):
        self.source_dir = Path(source_dir)
        self.target_dir = Path(target_dir)
        self.target_dir.mkdir(parents=True, exist_ok=True)

    def extract_department_from_filename(self, filename: str) -> Tuple[str, str]:
        """
        Extract department code and responsible person from filename
        Pattern: КПЭ 2025_DEPT (NAME)+.xlsx

        Returns: (dept_code, responsible_name)
        """
        match = re.match(r"КПЭ 2025_(.+?)(?:\s*\((.+?)\))?[+_]?\.xlsx?$", filename)
        if match:
            dept = match.group(1).strip()
            employee = match.group(2).strip() if match.group(2) else ""
            return dept, employee
        return filename.replace(".xlsx", "").replace(".xls", ""), ""

    def extract_positions_and_names(self, excel_path: Path) -> Dict[str, str]:
        """
        Extract position titles and employee names from Excel headers

        Returns: {
            "Директор по информационным технологиям": "Сложеникин А.В.",
            "Руководитель отдела": "Нор Е.А."
        }
        """
        wb = load_workbook(excel_path, data_only=True)

        # Find main sheet (usually "Заполнить")
        main_sheet_name = "Заполнить" if "Заполнить" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[main_sheet_name]

        # Get rows 2 and 3 (1-indexed in Excel = 0-indexed here + 1)
        rows = list(ws.iter_rows(values_only=True, min_row=2, max_row=3))

        if len(rows) < 2:
            return {}

        row_positions = rows[0]  # Row 2: Position titles
        row_names = rows[1]      # Row 3: Employee names

        positions_map = {}
        position_count = {}  # Track duplicate positions

        for pos, name in zip(row_positions, row_names):
            if not pos or not name:
                continue

            pos_str = str(pos).strip()
            name_str = str(name).strip()

            # Skip non-position columns
            if pos_str in ['КПЭ', '']:
                continue

            # Handle duplicate position titles (e.g., "Руководитель управления" x3)
            if pos_str in position_count:
                position_count[pos_str] += 1
                # Add number suffix for duplicates
                unique_pos = f"{pos_str} (позиция {position_count[pos_str]})"
            else:
                position_count[pos_str] = 1
                unique_pos = pos_str

            positions_map[unique_pos] = name_str

        return positions_map

    def clarify_column_name(self, position_title: str, employee_name: str) -> str:
        """
        Create clarified column name for MD table

        Input: "Руководитель управления", "Дубровин Александр Сергеевич"
        Output: "Рук. управления (Дубровин)"
        """
        # Abbreviate common words
        abbr_map = {
            "Директор по информационным технологиям": "Директор по ИТ",
            "Руководитель отдела": "Рук. отдела",
            "Руководитель управления": "Рук. управления",
            "Заместитель": "Зам.",
        }

        short_title = position_title
        for full, abbr in abbr_map.items():
            if full in position_title:
                short_title = abbr
                break

        # Extract last name from full name
        name_parts = employee_name.split()
        last_name = name_parts[0] if name_parts else employee_name[:15]

        return f"{short_title} ({last_name})"

    def read_kpi_data(self, excel_path: Path) -> pd.DataFrame:
        """Read KPI data from Excel file"""
        # Read with row 2 as header (0-indexed = row 1)
        df = pd.read_excel(excel_path, sheet_name='Заполнить', header=1)

        # First row after header contains actual column names (КПЭ, Целевое значение, etc.)
        actual_columns = df.iloc[0].values

        # Create new column names combining position titles with actual column names
        new_columns = []
        for i, col in enumerate(df.columns):
            if i < len(actual_columns) and pd.notna(actual_columns[i]):
                new_columns.append(str(actual_columns[i]))
            else:
                new_columns.append(col)

        df.columns = new_columns

        # Drop the first row (it was the column names)
        df = df.iloc[1:].reset_index(drop=True)

        return df

    def generate_md_table(
        self,
        df: pd.DataFrame,
        positions_map: Dict[str, str],
        dept_code: str
    ) -> str:
        """
        Generate markdown table with clarified column names

        Strategy:
        1. Keep first 3 columns as-is (КПЭ, Целевое значение, Ед. изм.)
        2. Replace position columns with clarified names
        3. Keep metadata columns (Тип, Min, Max, Методика, Источник)
        """
        # Clean the dataframe
        df = df.copy()

        # Remove completely empty rows
        df = df.dropna(how='all')

        # Replace NaN with "-" for better readability
        df = df.fillna("-")

        # Build clarified column names
        clarified_columns = []
        position_index = 0
        position_keys = list(positions_map.keys())

        for col in df.columns:
            col_str = str(col)

            # Check if this is a position column
            if any(pos_title in col_str for pos_title in [
                "Директор", "Руководитель", "Заместитель", "Начальник"
            ]):
                # This is a position column - clarify it
                if position_index < len(position_keys):
                    pos_key = position_keys[position_index]
                    emp_name = positions_map[pos_key]
                    clarified_col = self.clarify_column_name(pos_key, emp_name)
                    clarified_columns.append(clarified_col)
                    position_index += 1
                else:
                    clarified_columns.append(col_str)
            else:
                # Keep as-is
                clarified_columns.append(col_str)

        df.columns = clarified_columns

        # Generate markdown table
        md_lines = []

        # Header row
        header = "| " + " | ".join(df.columns) + " |"
        md_lines.append(header)

        # Separator row
        separator = "| " + " | ".join([":---" for _ in df.columns]) + " |"
        md_lines.append(separator)

        # Data rows
        for _, row in df.iterrows():
            # Clean cell values
            cells = []
            for cell in row:
                cell_str = str(cell).strip()
                # Replace newlines with <br> for markdown
                cell_str = cell_str.replace("\n", "<br>")
                # Limit cell length for readability
                if len(cell_str) > 150:
                    cell_str = cell_str[:147] + "..."
                cells.append(cell_str)

            row_md = "| " + " | ".join(cells) + " |"
            md_lines.append(row_md)

        return "\n".join(md_lines)

    def convert_file(self, excel_path: Path) -> Optional[Path]:
        """
        Convert single Excel file to Hybrid MD format

        Returns: Path to created MD file, or None if failed
        """
        print(f"Converting {excel_path.name}...")

        try:
            # Extract metadata
            dept_code, responsible = self.extract_department_from_filename(excel_path.name)
            positions_map = self.extract_positions_and_names(excel_path)

            if not positions_map:
                print(f"  ⚠️  Warning: Could not extract positions map")
                # Continue anyway with basic conversion

            # Read data
            df = self.read_kpi_data(excel_path)

            # Generate YAML frontmatter
            metadata = {
                "department": dept_code,
                "responsible": responsible if responsible else "Не указан",
                "positions_map": positions_map,
                "source_file": excel_path.name,
                "format_version": "1.0",
                "description": f"KPI показатели для {dept_code}"
            }

            yaml_frontmatter = yaml.dump(
                metadata,
                allow_unicode=True,
                default_flow_style=False,
                sort_keys=False
            )

            # Generate MD table
            md_table = self.generate_md_table(df, positions_map, dept_code)

            # Combine
            content = f"---\n{yaml_frontmatter}---\n\n{md_table}\n"

            # Save to file
            # Clean dept_code for filename
            safe_dept_code = re.sub(r'[^\w\-]', '_', dept_code)
            output_filename = f"KPI_{safe_dept_code}.md"
            output_path = self.target_dir / output_filename

            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(content)

            print(f"  ✅ Created: {output_path.name}")
            print(f"  📊 Positions mapped: {len(positions_map)}")
            print(f"  📋 KPI rows: {len(df)}")

            return output_path

        except Exception as e:
            print(f"  ❌ Error: {e}")
            import traceback
            traceback.print_exc()
            return None

    def convert_all(self) -> List[Path]:
        """Convert all Excel files in source directory"""
        excel_files = list(self.source_dir.glob("КПЭ*.xlsx"))

        if not excel_files:
            print(f"❌ No Excel files found in {self.source_dir}")
            return []

        print(f"\n🔄 Found {len(excel_files)} Excel files to convert\n")

        converted_files = []
        for excel_file in sorted(excel_files):
            result = self.convert_file(excel_file)
            if result:
                converted_files.append(result)
            print()  # Empty line between files

        print(f"\n✅ Conversion completed!")
        print(f"📁 {len(converted_files)}/{len(excel_files)} files converted successfully")
        print(f"📂 Output directory: {self.target_dir}")

        return converted_files


def main():
    """Main entry point"""
    import argparse

    parser = argparse.ArgumentParser(
        description="Convert Excel KPI files to Hybrid MD format (YAML + MD table)"
    )
    parser.add_argument(
        "--source",
        default="/home/yan/A101/HR/docs/KPI",
        help="Source directory with Excel files"
    )
    parser.add_argument(
        "--target",
        default="/home/yan/A101/HR/data/KPI",
        help="Target directory for MD files"
    )

    args = parser.parse_args()

    converter = KPIToHybridMDConverter(args.source, args.target)
    converted_files = converter.convert_all()

    if converted_files:
        print("\n📄 Created files:")
        for file_path in converted_files:
            print(f"  - {file_path.name}")


if __name__ == "__main__":
    main()
